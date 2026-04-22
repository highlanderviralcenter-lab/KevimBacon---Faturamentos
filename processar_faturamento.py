#!/usr/bin/env python3
"""
processar_faturamento.py
========================
Automatiza o preenchimento da planilha de faturamento a partir das batidas de ponto.

LOGICA CENTRAL:
    - Conta quantos LOGINs cada funcionario tem no mes (= dias assinados na folha)
    - Compara com o maximo do grupo (PAR ou IMPAR) definido na aba MES
    - indisponibilidade = maximo_do_grupo - logins_registrados
    - disponibilidade   = logins_registrados
    - Os eventos FALTA/ATESTAD identificam datas e motivos especificos

USO:
    python processar_faturamento.py batidas.xlsx Faturamento.xlsx [saida.xlsx]
"""

import sys, os, shutil, math
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl
import pandas as pd


# ─── Utilitarios ────────────────────────────────────────────────────────────

def norm(texto):
    return str(texto or '').strip().lower()

def grupo_escala(escala_raw):
    s = str(escala_raw or '').upper()
    if 'ÍMPAR' in s or 'IMPAR' in s:
        return 'ÍMPAR'
    return 'PAR'

def para_datetime(d):
    if isinstance(d, datetime):
        return d
    return datetime.combine(d, datetime.min.time())


# ─── Leitura das batidas ─────────────────────────────────────────────────────

def ler_batidas(caminho):
    xl = pd.ExcelFile(caminho)
    frames = []
    for aba in xl.sheet_names:
        try:
            df = pd.read_excel(caminho, sheet_name=aba)
            if {'NOME', 'TIPO_EVENTO', 'DATA_HORA'}.issubset(df.columns):
                frames.append(df)
        except Exception:
            pass
    if not frames:
        raise ValueError("Nenhuma aba valida encontrada em batidas.xlsx")
    df = pd.concat(frames, ignore_index=True)
    df['DATA'] = pd.to_datetime(df['DATA_HORA']).dt.date
    return df


# ─── Leitura da aba MES ──────────────────────────────────────────────────────

MESES_PT = {
    'JANEIRO':1,'FEVEREIRO':2,'MARÇO':3,'ABRIL':4,'MAIO':5,'JUNHO':6,
    'JULHO':7,'AGOSTO':8,'SETEMBRO':9,'OUTUBRO':10,'NOVEMBRO':11,'DEZEMBRO':12,
}

def ler_tabela_mes(wb):
    ws = wb['MÊS']
    tabela = {}
    for row in ws.iter_rows(min_row=2, max_row=14, values_only=True):
        nome_mes = str(row[0] or '').upper().strip()
        if nome_mes in MESES_PT:
            num = MESES_PT[nome_mes]
            ini = row[1].date() if isinstance(row[1], datetime) else None
            fim = row[2].date() if isinstance(row[2], datetime) else None
            tabela[num] = {'PAR': row[7], 'ÍMPAR': row[8], 'inicio': ini, 'fim': fim}
    return tabela


# ─── Leitura da aba CARGO ────────────────────────────────────────────────────

def ler_cargo(wb):
    ws = wb['CARGO']
    por_nome = {}
    for row in ws.iter_rows(min_row=3, max_row=528, values_only=True):
        if row[2] is None:
            continue
        nome     = str(row[2]).strip()
        matr     = row[1]
        admissao = row[7].date() if isinstance(row[7], datetime) else None
        por_nome[norm(nome)] = {'matricula': matr, 'admissao': admissao, 'nome': nome}
    return por_nome


# ─── Analise das batidas ─────────────────────────────────────────────────────

def analisar_batidas(df_bat):
    """
    Por funcionario, conta LOGINs (dias trabalhados) e lista eventos FALTA/ATESTAD.
    Retorna: { nome_norm: { logins, ausencias: [{data, motivo}], nome_original } }
    """
    resultado = {}
    for nome, grp in df_bat.groupby('NOME'):
        nome_n = norm(nome)
        logins = grp[grp['TIPO_EVENTO'] == 'LOGIN']['DATA'].nunique()
        ausencias = []
        for _, row in grp[grp['TIPO_EVENTO'].isin(['FALTA', 'ATESTAD'])].iterrows():
            motivo = 'Atestado Medico' if row['TIPO_EVENTO'] == 'ATESTAD' else 'Falta Injustificada'
            ausencias.append({'data': row['DATA'], 'motivo': motivo})
        ausencias.sort(key=lambda x: x['data'])
        resultado[nome_n] = {
            'logins': logins,
            'ausencias': ausencias,
            'nome_original': str(nome).strip(),
        }
    return resultado


# ─── Maximo ajustado para admissoes parciais ─────────────────────────────────

def max_ajustado(max_grupo, admissao, info_mes):
    if admissao is None or info_mes is None or info_mes.get('inicio') is None:
        return max_grupo
    if admissao <= info_mes['inicio']:
        return max_grupo
    dias_corridos = (info_mes['fim'] - admissao).days + 1
    total_mes     = (info_mes['fim'] - info_mes['inicio']).days + 1
    return math.floor(max_grupo * dias_corridos / total_mes)


# ─── Agrupar ausencias em periodos contiguos ─────────────────────────────────

def agrupar_em_periodos(ausencias_lista):
    if not ausencias_lista:
        return []
    periodos, inicio, fim, motivos = [], ausencias_lista[0]['data'], ausencias_lista[0]['data'], [ausencias_lista[0]['motivo']]
    for item in ausencias_lista[1:]:
        if (item['data'] - fim).days <= 2:
            fim = item['data']
            motivos.append(item['motivo'])
        else:
            periodos.append(_periodo(inicio, fim, motivos))
            inicio, fim, motivos = item['data'], item['data'], [item['motivo']]
    periodos.append(_periodo(inicio, fim, motivos))
    return periodos

def _periodo(inicio, fim, motivos):
    motivo = 'Atestado Medico' if 'Atestado Medico' in motivos else 'Falta Injustificada'
    return {'inicio': inicio, 'fim': fim, 'total': len(motivos), 'motivo': motivo, 'retorno': fim + timedelta(days=2)}


# ─── Colunas FATURAMENTO (base 1) ────────────────────────────────────────────

COL_NOME       = 4   # D
COL_ESCALA     = 12  # L
COL_INDISP_INI = 16  # P
COL_INDISP_FIM = 17  # Q
COL_TOT_INDISP = 19  # S
COL_TOT_DISP   = 20  # T


# ─── Funcao principal ────────────────────────────────────────────────────────

def processar(batidas_path, faturamento_path, output_path):
    sep = '=' * 60
    print(sep)
    print('  PROCESSAMENTO DE FATURAMENTO')
    print(sep)

    print(f'\n[1/5] Lendo batidas: {batidas_path}')
    df_bat  = ler_batidas(batidas_path)
    mes_num = int(pd.to_datetime(df_bat['DATA_HORA']).dt.month.mode()[0])
    print(f'      {len(df_bat):,} registros | mes detectado: {mes_num}')

    print('\n[2/5] Analisando logins e ausencias...')
    analise = analisar_batidas(df_bat)
    print(f'      {len(analise)} funcionario(s) nas batidas.')

    print(f'\n[3/5] Abrindo planilha: {faturamento_path}')
    shutil.copy2(faturamento_path, output_path)
    wb_ro = openpyxl.load_workbook(output_path, data_only=True)
    wb    = openpyxl.load_workbook(output_path)

    tabela_mes = ler_tabela_mes(wb_ro)
    info_mes   = tabela_mes.get(mes_num, {'PAR': 15, 'ÍMPAR': 16, 'inicio': None, 'fim': None})
    dias_grupo = {'PAR': info_mes['PAR'], 'ÍMPAR': info_mes['ÍMPAR']}
    print(f'      PAR={dias_grupo["PAR"]} dias | ÍMPAR={dias_grupo["ÍMPAR"]} dias')

    cargo_por_nome = ler_cargo(wb_ro)

    # Mapear nome → linha e escala no FATURAMENTO
    ws_ro   = wb_ro['FATURAMENTO']
    ws_fat  = wb['FATURAMENTO']
    linhas  = {}
    escalas = {}
    for row in ws_ro.iter_rows(min_row=5, max_row=667):
        nc = row[COL_NOME - 1]
        ec = row[COL_ESCALA - 1]
        if nc.value:
            nn = norm(nc.value)
            linhas[nn]  = nc.row
            escalas[nn] = grupo_escala(ec.value)

    print('\n[4/5] Atualizando FATURAMENTO...')
    avisos = []

    for nome_norm, linha in linhas.items():
        grupo    = escalas.get(nome_norm, 'PAR')
        max_std  = dias_grupo[grupo]
        admissao = cargo_por_nome.get(nome_norm, {}).get('admissao')
        max_dias = max_ajustado(max_std, admissao, info_mes)

        bat = analise.get(nome_norm)

        if bat is None:
            avisos.append(f'  NAO nas batidas: {nome_norm}')
            ws_fat.cell(row=linha, column=COL_TOT_DISP).value = 0
            ws_fat.cell(row=linha, column=COL_TOT_INDISP).value = max_dias
            ws_fat.cell(row=linha, column=COL_INDISP_INI).value = None
            ws_fat.cell(row=linha, column=COL_INDISP_FIM).value = None
            continue

        logins       = bat['logins']
        # ── REGRA CENTRAL ─────────────────────────────────────────
        # disp = logins registrados
        # indisp = max_grupo - logins
        # disp + indisp = max_grupo  (SEMPRE, sem excecao)
        # A chave e o GRUPO — se o grupo estiver errado na coluna L,
        # toda a conta fica errada. O script confia no grupo da planilha.
        # ──────────────────────────────────────────────────────────

        disponib     = logins
        total_indisp = max_dias - logins   # pode ser negativo se grupo errado

        # ERRO CRITICO: logins excedem o maximo do grupo
        # Isso so acontece se o grupo estiver errado na planilha (coluna L)
        # Ex: funcionario IMPAR com 16 logins classificado como PAR (max 15)
        if total_indisp < 0:
            avisos.append(
                f'  !! GRUPO ERRADO? {nome_norm}: {logins} logins EXCEDE '
                f'max {max_dias} [{grupo}] — verificar coluna ESCALA no FATURAMENTO'
            )
            # Corrige para nao salvar valor negativo, mas mantem o aviso
            total_indisp = 0
            disponib     = logins

        # Garantia matematica: disp + indisp deve sempre fechar o mes
        assert disponib + total_indisp == max_dias or total_indisp == 0, \
            f'FALHA na equacao para {nome_norm}: {disponib}+{total_indisp} != {max_dias}'

        # Aviso: ausencias sem motivo registrado no ponto eletronico
        # (atestados em papel, por exemplo, nao aparecem nas batidas)
        n_reg = len(bat['ausencias'])
        if total_indisp > 0 and total_indisp > n_reg:
            diff = total_indisp - n_reg
            avisos.append(
                f'  SEM MOTIVO: {nome_norm} — {total_indisp} dia(s) indisp. '
                f'mas so {n_reg} evento(s) no ponto ({diff} sem justificativa registrada)'
            )

        # Datas (primeiro e ultimo evento de ausencia registrado)
        ini_i = bat['ausencias'][0]['data']  if bat['ausencias'] else None
        fim_i = bat['ausencias'][-1]['data'] if bat['ausencias'] else None

        ws_fat.cell(row=linha, column=COL_TOT_DISP).value = disponib
        if total_indisp > 0:
            ws_fat.cell(row=linha, column=COL_TOT_INDISP).value = total_indisp
            ws_fat.cell(row=linha, column=COL_INDISP_INI).value = para_datetime(ini_i) if ini_i else None
            ws_fat.cell(row=linha, column=COL_INDISP_FIM).value = para_datetime(fim_i) if fim_i else None
            print(f'      FALTA {nome_norm}: {logins} login(s) | indisp={total_indisp} | disp={disponib}+indisp={total_indisp}={disponib+total_indisp} [{grupo} max={max_dias}]')
        else:
            ws_fat.cell(row=linha, column=COL_TOT_INDISP).value = None
            ws_fat.cell(row=linha, column=COL_INDISP_INI).value = None
            ws_fat.cell(row=linha, column=COL_INDISP_FIM).value = None
            print(f'      OK    {nome_norm}: {logins} login(s) | disp={disponib}+indisp=0={disponib} [{grupo} max={max_dias}]')

    print('\n[5/5] Preenchendo INDISPONIBILIDADE...')
    total_reg = _preencher_indisponibilidade(wb, analise, cargo_por_nome)
    print(f'      {total_reg} registro(s) inserido(s).')

    wb.save(output_path)

    print(f'\n{sep}')
    print(f'  CONCLUIDO: {output_path}')
    print(sep)

    if avisos:
        print('\nAVISOS (verificar manualmente):')
        for a in avisos:
            print(a)
        print()


def _preencher_indisponibilidade(wb, analise, cargo_por_nome):
    ws = wb['INDISPONIBILIDADE']
    # Limpar dados anteriores (mantém cabeçalho linha 2)
    for row in ws.iter_rows(min_row=3, max_row=200):
        for cell in row:
            cell.value = None
    linha = 3
    for nome_norm, bat in analise.items():
        if not bat['ausencias']:
            continue
        matricula   = cargo_por_nome.get(nome_norm, {}).get('matricula', '')
        nome_exibir = bat['nome_original']
        for p in agrupar_em_periodos(bat['ausencias']):
            ws.cell(row=linha, column=2).value = matricula
            ws.cell(row=linha, column=3).value = nome_exibir
            ws.cell(row=linha, column=4).value = p['motivo']
            ws.cell(row=linha, column=5).value = para_datetime(p['inicio'])
            ws.cell(row=linha, column=6).value = para_datetime(p['fim'])
            ws.cell(row=linha, column=7).value = p['total']
            ws.cell(row=linha, column=8).value = para_datetime(p['retorno'])
            linha += 1
    return linha - 3


# ─── Entrypoint ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    args = sys.argv[1:]
    if len(args) < 2:
        print(__doc__)
        sys.exit(1)
    batidas_path     = args[0]
    faturamento_path = args[1]
    saida_path       = args[2] if len(args) > 2 else 'Faturamento_PROCESSADO.xlsx'
    for f in [batidas_path, faturamento_path]:
        if not os.path.exists(f):
            print(f'ERRO: arquivo nao encontrado: {f}')
            sys.exit(1)
    processar(batidas_path, faturamento_path, saida_path)
